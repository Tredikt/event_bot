from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

from core.utils.enums import Variables


class AnalyticsService:
    
    def __init__(self):
        # Маппинг названий интерактивов к колонкам Excel
        self.columns_mapping = {
            # Белозёрцева
            "belozyortseva_start_interactive": "Запуск интерактив Белозёрцевой (кнопка - запустить интерактив)",
            "belozyortseva_question_1": "Первый вопрос (1 балл)",
            "belozyortseva_question_2": "Второй вопрос (1 балл)", 
            "belozertseva_cool": "Круто, понравилось! (1 балл)",
            "belozertseva_notbad": "Неплохо (1 балл)",
            "belozertseva_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Мендубаев
            "mendubaev_cool": "Мендубаев - Круто, понравилось! (1 балл)",
            "mendubaev_notbad": "Мендубаев - Неплохо (1 балл)", 
            "mendubaev_ending": "Мендубаев - Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Садриев
            "sadriev_start_interactive": "Запуск интерактив Садриев (кнопка - запустить интерактив)",
            "sadriev_question_1": "Вопрос (1 балл)",
            "sadriev_cool": "Круто, понравилось! (1 балл)",
            "sadriev_notbad": "Неплохо (1 балл)",
            "sadriev_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Адмакин
            "admakin_cool": "Адмакин - Круто, понравилось! (1 балл)",
            "admakin_notbad": "Адмакин - Неплохо (1 балл)",
            "admakin_ending": "Адмакин - Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Нурхаметова
            "nurkhametova_start_interactive": "Запуск интерактив Нурхаметова (кнопка - запустить интерактив)",
            "nurkhametova_question_1": "Вопрос 1 (1 балл)",
            "nurkhametova_question_2": "Вопрос 2 (1 балл)",
            "nurhametova_cool": "Круто, понравилось! (1 балл)",
            "nurhametova_notbad": "Неплохо (1 балл)",
            "nurhametova_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Балуков
            "balukov_cool": "Балуков - Круто, понравилось! (1 балл)",
            "balukov_notbad": "Балуков - Неплохо (1 балл)",
            "balukov_ending": "Балуков - Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Хорошутина
            "horoshutina_start_interactive": "Запуск интерактив Хорошутина (кнопка - запустить интерактив)",
            "horoshutina_interactive": "Собрал цепочку (2 балла)",
            "horoshutina_cool": "Круто, понравилось! (1 балл)",
            "horoshutina_notbad": "Неплохо (1 балл)",
            "horoshutina_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Забегаев
            "zabegayev_start_interactive": "Запуск интерактив Забегаев (кнопка - запустить интерактив)",
            "zabegayev_question_1": "Вопрос 1 (1 балл)",
            "zabegayev_question_2": "Вопрос 2 (1 балл)",
            "zabegaev_cool": "Круто, понравилось! (1 балл)",
            "zabegaev_notbad": "Неплохо (1 балл)",
            "zabegaev_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Гавриков
            "gavrikov_start_interactive": "Запуск интерактив Гавриков (кнопка - запустить интерактив)",
            "gavrikov_choice": "Выбрал вариант (1 балл)",
            "gavrikov_cool": "Круто, понравилось! (1 балл)",
            "gavrikov_notbad": "Неплохо (1 балл)",
            "gavrikov_ending": "Оставил инсайт\\ задал вопрос (2 балла)",
            
            # Гильманова
            "gilmanova_cool": "Гильманова - Круто, понравилось! (1 балл)",
            "gilmanova_notbad": "Гильманова - Неплохо (1 балл)",
            "gilmanova_ending": "Гильманова - Оставил инсайт\\ задал вопрос (2 балла)"
        }
        
        # Колонки с текстами инсайтов
        self.insight_text_columns = {
            "belozertseva": "Текст инсайта\\вопроса",
            "mendubaev": "Текст инсайта\\вопроса",  
            "sadriev": "Текст инсайта\\вопроса",
            "admakin": "Текст инсайта\\вопроса",
            "nurhametova": "Текст инсайта\\вопроса",
            "balukov": "Текст инсайта\\вопроса",
            "horoshutina": "Текст инсайта\\вопроса",
            "zabegaev": "Текст инсайта\\вопроса",
            "gavrikov": "Текст инсайта\\вопроса",
            "gilmanova": "Текст инсайта\\вопроса"
        }

    async def generate_analytics_excel(self, variables: Variables) -> bytes:
        """Генерирует Excel файл с аналитикой"""
        
        # Получаем все данные
        users = await variables.db.user.get_all_users()
        interactive_history = await self._get_all_interactive_history(variables)
        feedback_data = await self._get_all_feedback(variables)
        
        # Группируем данные по пользователям
        user_data = await self._process_user_data(users, interactive_history, feedback_data)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Аналитика ТЕХНОБАРС"
        
        # Создаем заголовки
        headers = self._create_headers()
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Устанавливаем высоту строки заголовков
        ws.row_dimensions[1].height = 40
        
        # Сортируем пользователей по общей сумме баллов
        sorted_users = sorted(user_data.items(), key=lambda x: x[1]['total_points'], reverse=True)
        
        # Заполняем данные
        for row, (user_id, data) in enumerate(sorted_users, 2):
            self._fill_user_row(ws, row, user_id, data, sorted_users)
        
        # Автоширина колонок
        headers = self._create_headers()
        for col_index, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = column[0].column_letter
            
            # Проверяем, является ли это колонкой с текстом инсайта
            is_insight_column = False
            if col_index <= len(headers):
                header = headers[col_index - 1]
                if "Текст инсайта" in header:
                    is_insight_column = True
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Для колонок с инсайтами устанавливаем минимальную ширину 40, максимальную 80
            if is_insight_column:
                adjusted_width = max(min(max_length + 2, 80), 40)
            else:
                adjusted_width = min(max_length + 2, 50)
                
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в память
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()

    async def _get_all_interactive_history(self, variables: Variables) -> List[Dict]:
        """Получает всю историю интерактивов"""
        histories = await variables.db.interactive_history.get_all_history()
        return [
            {
                'user_id': h.user_id,
                'interactive_name': h.interactive_name,
                'points_earned': h.points_earned,
                'completed_at': h.completed_at
            }
            for h in histories
        ]

    async def _get_all_feedback(self, variables: Variables) -> List[Dict]:
        """Получает все отзывы и инсайты"""
        feedbacks = await variables.db.feedback.get_all_feedback()
        return [
            {
                'user_id': f.user_id,
                'interactive_name': f.interactive_name,
                'rate': f.rate,
                'inside': f.inside
            }
            for f in feedbacks
        ]

    async def _process_user_data(self, users, interactive_history, feedback_data) -> Dict[str, Dict]:
        """Обрабатывает данные пользователей"""
        user_data = {}
        
        # Инициализируем данные пользователей
        for user in users:
            user_data[user.user_id] = {
                'username': user.username or user.first_name or f"User_{user.user_id[:6]}",
                'total_points': 0,
                'activities': {},
                'insights': {}
            }
        
        # Обрабатываем историю интерактивов
        for history in interactive_history:
            user_id = history['user_id']
            if user_id in user_data:
                activity_name = history['interactive_name']
                points = history['points_earned']
                
                user_data[user_id]['total_points'] += points
                
                if activity_name in user_data[user_id]['activities']:
                    user_data[user_id]['activities'][activity_name] += points
                else:
                    user_data[user_id]['activities'][activity_name] = points
        
        # Обрабатываем отзывы и инсайты
        for feedback in feedback_data:
            user_id = feedback['user_id']
            if user_id in user_data:
                interactive_name = feedback['interactive_name']
                rate = feedback['rate']
                inside = feedback['inside']
                
                # Определяем тип активности по rate
                if rate == "Круто":
                    activity_name = f"{interactive_name}_cool"
                elif rate == "Неплохо":
                    activity_name = f"{interactive_name}_notbad"
                elif rate == "insight":
                    activity_name = f"{interactive_name}_ending"
                    # Сохраняем текст инсайта с ключом interactive_name
                    if inside:  # Сохраняем только если есть текст
                        user_data[user_id]['insights'][interactive_name] = inside
                else:
                    continue
                
                # Добавляем активность если её еще нет
                if activity_name not in user_data[user_id]['activities']:
                    user_data[user_id]['activities'][activity_name] = 1
                    user_data[user_id]['total_points'] += 1
        
        return user_data

    def _calculate_cell_height(self, text: str, column_width: float) -> float:
        """Вычисляет необходимую высоту ячейки для текста с переносом"""
        if not text:
            return 15
        
        # Учитываем ширину колонки (в Excel примерно 7 пикселей на символ)
        chars_per_line = max(int(column_width * 0.7), 15)
        
        # Разбиваем текст на слова и подсчитываем строки
        words = text.split()
        lines = 1
        current_line_length = 0
        
        for word in words:
            # Если слово не помещается в текущую строку
            if current_line_length + len(word) + 1 > chars_per_line:
                lines += 1
                current_line_length = len(word)
            else:
                current_line_length += len(word) + (1 if current_line_length > 0 else 0)
        
        # Учитываем принудительные переносы строк
        lines += text.count('\n')
        
        # Примерно 15 пикселей на строку + отступы
        return min(lines * 15 + 10, 120)

    def _create_headers(self) -> List[str]:
        """Создает заголовки для Excel"""
        headers = [
            "ник сотрудника (/start):",
            "СУММА:",
            "Распределение от большего к меньшему (по сумме) - ник:",
            "Распределение от большего к меньшему (по сумме) - сумма:"
        ]
        
        # Добавляем колонки для каждого интерактива
        for activity_key, column_name in self.columns_mapping.items():
            headers.append(column_name)
            
            # Добавляем колонку с текстом инсайта после активностей "ending"
            if "_ending" in activity_key:
                # Извлекаем имя спикера из ключа активности
                speaker_name = activity_key.replace("_ending", "")
                if speaker_name in self.insight_text_columns:
                    headers.append(self.insight_text_columns[speaker_name])
        
        return headers

    def _fill_user_row(self, ws, row: int, user_id: str, data: Dict, sorted_users: List):
        """Заполняет строку пользователя"""
        col = 1
        
        # Ник пользователя
        ws.cell(row=row, column=col, value=data['username'])
        col += 1
        
        # Общая сумма
        ws.cell(row=row, column=col, value=data['total_points'])
        col += 1
        
        # Рейтинг (ник)
        rank = next(i for i, (uid, _) in enumerate(sorted_users, 1) if uid == user_id)
        ws.cell(row=row, column=col, value=data['username'])
        col += 1
        
        # Рейтинг (сумма)
        ws.cell(row=row, column=col, value=data['total_points'])
        col += 1
        
        # Переменная для отслеживания максимальной высоты строки
        max_height = 15  # Минимальная высота строки
        
        # Активности
        for activity_key, column_name in self.columns_mapping.items():
            value = data['activities'].get(activity_key, 0)
            ws.cell(row=row, column=col, value=value if value > 0 else "")
            col += 1
            
            # Добавляем текст инсайта если это активность "_ending"
            if "_ending" in activity_key:
                # Извлекаем имя спикера из ключа активности
                speaker_name = activity_key.replace("_ending", "")
                if speaker_name in self.insight_text_columns:
                    insight_text = data['insights'].get(speaker_name, "")
                    cell = ws.cell(row=row, column=col, value=insight_text)
                    
                    # Добавляем перенос текста для ячеек с инсайтами
                    if insight_text:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Вычисляем необходимую высоту для этой ячейки
                        column_width = ws.column_dimensions[cell.column_letter].width or 40
                        cell_height = self._calculate_cell_height(insight_text, column_width)
                        
                        max_height = max(max_height, cell_height)
                    
                    col += 1
        
        # Устанавливаем высоту строки на основе самой высокой ячейки
        ws.row_dimensions[row].height = max_height 